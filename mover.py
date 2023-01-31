import openpyxl
import pathlib
import shutil
import click
import sys
import re
import os


# TODO : Add GUI, I think I'd like to use PySimpleGUI for this.
# TODO : Add overwrite protection?
# TODO : Add time stamp to the log file.
def enum_excel_rows(excel_file: str, sheet, no_header) -> dict:
    """
    Extracts filepaths from the given Excel file
    Args:
        excel_file: The filepath index (Excel) file
        sheet: The sheet to operate on
        no_header: If the sheet has a header or not (Column names)

    Returns:

    """
    file = pathlib.Path(excel_file)
    if file.is_file():
        file_info = {}
        click.echo(click.style("Success", fg="green") + " - " + "Source file found.")
        workbook = openpyxl.load_workbook(file)
        ws = workbook[sheet]

        min_row = ws.min_row if no_header else ws.min_row + 1

        for row_cells in ws.iter_rows(min_row=min_row, max_row=ws.max_row, max_col=3):
            regex = re.compile(r"^\D*\.(\D)(\d*)>$")
            matches = regex.findall(str(row_cells[0]))[0]
            row = matches[1]
            file_info[row] = {'file': row_cells[0].value, "source": row_cells[1].value, "dest": row_cells[2].value}
        return file_info
    elif file.is_dir():
        # TODO : I could add folder handling. When a folder is specified it enumerates the files in that dir instead of getting files form Excel.
        click.echo(click.style("Error", fg="red") + " - " + "Given path is a folder, not an Excel file." + click.style(" Aborting!", fg="red"))
    else:
        click.echo(click.style("Error", fg="red") + " - " + "Source file not found." + click.style(" Aborting!", fg="red"))
    sys.exit()


def enum_files(files: dict, abs_path, move, test) -> dict:
    """
    Enumerates over a list of files. It will then copy or move the files.
    Args:
        files: A dictionary containing file info (as filename, source and destination)
        abs_path: If the paths given are absolute or not (default: 'False')
        move: If to move or copy the files (default: 'False'
        test: If you want to performa dry run (default: 'False'"

    Returns:
        A dict containing two lists. Status regarding the copy/move succeeded or were skipped.
    """
    status = {"skipped_files": [], "success": []}
    for row, info in files.items():
        if not abs_path:
            source_folder = pathlib.Path(os.getcwd(), info["source"])
            destination_folder = pathlib.Path(os.getcwd(), info["dest"])
        else:
            source_folder = info["source"]
            destination_folder = info["dest"]

        source_file = pathlib.Path(source_folder, info["file"])
        destination_file = pathlib.Path(destination_folder, info["file"])
        try:
            if source_file.is_file():
                if not destination_folder.exists():
                    destination_folder.mkdir(parents=True, exist_ok=True)
                if not move and not test:
                    shutil.copy2(source_file, destination_file)
                elif move and not test:
                    shutil.move(source_file, destination_file)
                click.echo(click.style("Success", fg="green") + " - " + f"{'Copied' if not move else 'Moved'} file ({click.style(source_file.name, fg='yellow')}) from '{click.style(source_folder, fg='magenta')}' to '{click.style(destination_folder, fg='cyan')}'")
                status["success"].append(str(source_file.resolve()))
            else:
                status["skipped_files"].append(str(source_file.resolve()))
        except:
            status["skipped_files"].append(str(source_file.resolve()))
    return status



@click.command()
@click.option("-s", "--src-file",
              help="Source excel file. If the path contains spaces, please surround them with quotes.",
              required=True)
@click.option("-ap", "--abs-path",
              help="Switches path handling to 'Absolute'.",
              is_flag=True,
              default=False)
@click.option("-sh", "--sheet",
              help="Specify the sheet to read from.",
              default="Sheet1")
@click.option("-m", "--move",
              help="Set this flag to move the files instead of copying.",
              is_flag=True,
              default=False)
@click.option("-t", '--test',
              help="Set this flag for a test run.",
              is_flag=True,
              default=False)
@click.option("-nh", '--no-header',
              help="Set this flag if your Excel files has no header.",
              is_flag=True,
              default=False)
def main(src_file, abs_path, sheet, no_header, move, test):
    """
    This is a small script to mass-copy files from one directory to another.
    Important!! The paths in the Excel file can be either relative (default) or absolute.
    If you choose to work with relative paths, please start this script IN the parent folder.


        E.g: If you want to move files:


            From: C:\\Users\\kbergene\\Documents\\SCO-ILD\\10.1

            To: C:\\Users\\kbergene\\Documents\\SCO-ILD_Load\\10.1

        Then you should run this script from the 'Documents' folder.


    It will create the destination folders, if needed.
    """
    files = enum_excel_rows(excel_file=src_file, sheet=sheet, no_header=no_header)
    status = enum_files(files=files, abs_path=abs_path, move=move, test=test)

    for file in status["skipped_files"]:
        click.echo(click.style('Skipped', fg='red') + " - " + click.style(file, fg='yellow'))
    click.echo(f"\n{click.style('Skipped', fg='yellow')}: {len(status['skipped_files'])}\n{click.style('Succeeded', fg='green')}: {len(status['success'])}\n{click.style('Total', fg='blue')}: {len(status['skipped_files']) + len(status['success'])}")

    # Log
    with open("log.txt", "a") as log:
        for key, value in status.items():
            for file in value:
                if key == "skipped_files": key = "Skipped"
                log.write(f"Status: {key.capitalize()} - File: {file}\n")
        log.write(f"\nSkipped: {len(status['skipped_files'])}\nSuccess: {len(status['success'])}\nTotal: {len(status['skipped_files']) + len(status['success'])}\n\n")

if __name__ == "__main__":
    main()
