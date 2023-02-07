import openpyxl
import openpyxl.utils.exceptions
import pathlib
import shutil
import click
import sys
import re
import os

"""
Author: Kåre Bergene
"""

# TODO : Add GUI, I think I'd like to use PySimpleGUI for this.
# TODO : Add overwrite protection?
# TODO : Add time stamp to the log filename and in the logged text.
# TODO : Log verbose and non-verbose text. Currently only some verbose text is logged.
# TODO : Implement Logging module.


def enum_excel_rows(excel_file: str, sheet, no_header, verbose) -> dict:
    """
    Extracts filepaths from the given Excel file
    Args:
        excel_file: The filepath index (Excel) file
        sheet: The sheet to operate on
        no_header: If the sheet has a header or not (Column names)

    Returns:

    """
    file = pathlib.Path(excel_file)
    if file.is_file():
        file_info = {}
        click.echo(click.style("Success", fg="green") + " - " + "Source file found.")
        try:
            workbook = openpyxl.load_workbook(file)
        except openpyxl.utils.exceptions.InvalidFileException as e:
            # TODO : Log this error
            click.echo(click.style("Error", fg="red") + " - " + "Invalid file extension. Supported formats are: .xlsx, .xlsm, .xltx, .xltm")
            sys.exit()
        ws = workbook[sheet]

        if no_header:
            min_row = ws.min_row
            click.echo(click.style("Working", fg="green") + " - " + "Header flag disabled..")
        else:
            min_row = ws.min_row + 1
            click.echo(click.style("Working", fg="green") + " - " + "Header flag enabled.")

        click.echo(click.style("Working", fg="green") + " - " + "Enumerating rows.")
        for row_cells in ws.iter_rows(min_row=min_row, max_row=ws.max_row, max_col=3):
            col_row = str(row_cells[0]).split(".")[-1]
            regex = re.compile(r"(\d*)>$")
            matches = regex.findall(col_row)
            if isinstance(matches[0], tuple):
                row = matches[0][1]
            else:
                row = matches[0]
            file_info[row] = {'file': row_cells[0].value, "source": row_cells[1].value, "dest": row_cells[2].value}
        click.echo(click.style("Working", fg="green") + " - " + "Enumeration done.")
        return file_info
    elif file.is_dir():
        # TODO : I could add folder handling. When a folder is specified it enumerates the files in that dir instead of getting files form Excel.
        click.echo(click.style("Error", fg="red") + " - " + "Given path is a folder, not an Excel file." + click.style(" Aborting!", fg="red"))
    else:
        click.echo(click.style("Error", fg="red") + " - " + "Source file not found." + click.style(" Aborting!", fg="red"))
    sys.exit()


def enum_files(files: dict, abs_path, move, test, verbose) -> dict:
    """
    Enumerates over a list of files. It will then copy or move the files.
    Args:
        files: A dictionary containing file info (as filename, source and destination)
        abs_path: If the paths given are absolute or not (default: 'False')
        move: If to move or copy the files (default: 'False'
        test: If you want to performa dry run (default: 'False'"

    Returns:
        A dict containing two lists. Status regarding the copy/move succeeded or were skipped.
    """

    click.echo(click.style("Working", fg="green") + " - " + f"{'Absolute' if abs_path else 'Relative'} paths enabled.")

    status = {"skipped_files": [], "success": []}

    click.echo(click.style("Working", fg="green") + " - " + f"{'Copying..' if not move else 'Moving..'}")

    for row, info in files.items():
        if not abs_path:
            source_folder = pathlib.Path(os.getcwd(), info["source"])
            destination_folder = pathlib.Path(os.getcwd(), info["dest"])
        else:
            source_folder = info["source"]
            destination_folder = info["dest"]

        source_file = pathlib.Path(source_folder, info["file"])
        destination_file = pathlib.Path(destination_folder, info["file"])
        try:
            if source_file.is_file():
                if verbose: click.echo(click.style("Success", fg="green") + " - " + "Source file exists.")
                if not destination_folder.exists() and not test:
                    destination_folder.mkdir(parents=True, exist_ok=True)
                    if verbose: click.echo(click.style("Success", fg="green") + " - " + f"Created directory: {destination_folder.resolve()}")
                if not move and not test:
                    shutil.copy2(source_file, destination_file)
                elif move and not test:
                    shutil.move(source_file, destination_file)
                if verbose: click.echo(click.style("Success", fg="green") + " - " + f"{'Copied' if not move else 'Moved'} file ({click.style(source_file.name, fg='yellow')}) from '{click.style(source_folder, fg='magenta')}' to '{click.style(destination_folder, fg='cyan')}'")
                #["success"].append(f"Excel row: {row} - Source: {str(source_file.resolve())}")
                status["success"].append({"row": row, "source": source_file.resolve()})
            else:
                if verbose: click.echo(
                    click.style("Warning", fg="yellow") + " - " + f"Skipped file: " + click.style(source_file.resolve(), fg="yellow"))
                #status["skipped_files"].append(f"Excel row: {row} - Source: {str(source_file.resolve())}")
                status["skipped_files"].append({"row": row, "source": source_file.resolve()})
        except:
            if verbose: click.echo(
                    click.style("Warning", fg="yellow") + " - " + f"Skipped file: " + click.style(source_file.resolve(), fg="yellow"))
            #status["skipped_files"].append(f"Excel row: {row} - Source: {str(source_file.resolve())}")
            status["skipped_files"].append({"row": row, "source": source_file.resolve()})

    click.echo(click.style("Completed", fg="green"))
    return status


@click.command()
@click.option("-s", "--src-file",
              help="Source excel file. If the path contains spaces, please surround them with quotes.",
              required=True)
@click.option("-ap", "--abs-path",
              help="Switches path handling to 'Absolute'.",
              is_flag=True,
              default=False)
@click.option("-sh", "--sheet",
              help="Specify the sheet to read from.",
              default="Sheet1")
@click.option("-m", "--move",
              help="Set this flag to move the files instead of copying.",
              is_flag=True,
              default=False)
@click.option("-t", '--test',
              help="Set this flag for a test run.",
              is_flag=True,
              default=False)
@click.option("-nh", '--no-header',
              help="Set this flag if your Excel files has no header.",
              is_flag=True,
              default=False)
@click.option("-v", "--verbose",
              help="Enables verbosity.",
              is_flag=True,
              default=False)
@click.option("-l", "--logging",
              help="Enables logging to file. File output is the current working directory.",
              is_flag=True,
              default=False)
def main(src_file, abs_path, sheet, no_header, move, test, verbose, logging):
    """
    This is a small script to mass-copy files from one directory to another.
    Important!! The paths in the Excel file can be either relative (default) or absolute.
    If you choose to work with relative paths, please start this script IN the parent folder.


        E.g: If you want to move files:


            From: C:\\Users\\kbergene\\Documents\\SCO-ILD\\10.1

            To: C:\\Users\\kbergene\\Documents\\SCO-ILD_Load\\10.1

        Then you should run this script from the 'Documents' folder.


    It will create the destination folders, if needed.
    """
    files = enum_excel_rows(excel_file=src_file, sheet=sheet, no_header=no_header, verbose=verbose)
    status = enum_files(files=files, abs_path=abs_path, move=move, test=test, verbose=verbose)

    if verbose:
        for file in status["skipped_files"]:
            click.echo(click.style('Skipped', fg='red') + " - Row:" + click.style(file["row"], fg='yellow') + " - Source:" + click.style(file["source"], fg='yellow'))
        for file in status["success"]:
            click.echo(click.style('Success', fg='green') + " - Row: " + click.style(file["row"], fg='blue') + " - Source: " + click.style(file["source"], fg='yellow'))
    click.echo(f"\n{click.style('Skipped', fg='yellow')}: {len(status['skipped_files'])}\n{click.style('Succeeded', fg='green')}: {len(status['success'])}\n{click.style('Total', fg='blue')}: {len(status['skipped_files']) + len(status['success'])}")

    # Log
    if logging:
        with open("log.txt", "a") as log:
            for key, value in status.items():
                for file in value:
                    if key == "skipped_files": key = "Skipped"
                    log.write(f"Status: {key.capitalize()} - Row: {file['row']} - File: {file['source']}\n")
            log.write(f"\nSkipped: {len(status['skipped_files'])}\nSuccess: {len(status['success'])}\nTotal: {len(status['skipped_files']) + len(status['success'])}\n\n")


if __name__ == "__main__":
    main()
